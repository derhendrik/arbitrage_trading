import openpyxl
from graphviz import Digraph
import math


class ArbitrageProject():
    def __init__(self):
        self.graph = Graph()

    def read_excel_file(self, filename):
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook["Tabelle1"]
        max_row = worksheet.max_row

        # Offset because relevant data of excel-file starts in row 3 / column 3
        offset = 3

        # Get currency names and create nodes
        for i in range(offset, max_row + 1):
            cell_entry = worksheet.cell(row=i, column=2)
            self.graph.set_of_nodes.append(Node(cell_entry.value))
            self.graph.table_header.append(cell_entry.value)

        self.graph.number_of_nodes = len(self.graph.set_of_nodes)
        self.graph.arc_matrix = [[0 for i in range(self.graph.number_of_nodes)] for j in
                                 range(self.graph.number_of_nodes)]

        # Get corresponding entries of trading pair matrix and create arcs
        for i, from_node in enumerate(self.graph.set_of_nodes):
            for j, to_node in enumerate(self.graph.set_of_nodes):
                cell_entry = worksheet.cell(row=i + offset, column=j + offset)
                if cell_entry.value != None:
                    arc_dummy = Arc(from_node, to_node, cell_entry.value)
                    self.graph.set_of_arcs.append(arc_dummy)
                    self.graph.arc_matrix[i][j] = arc_dummy

    def print_graph(self, filename):
        dot = Digraph(comment="Arbitrage-Graph")
        dot_log = Digraph(comment="Arbitrage-Graph_logarithmic")

        for node in self.graph.set_of_nodes:
            dot.node(node.name, node.name)
            dot_log.node(node.name, node.name)

        for arc in self.graph.set_of_arcs:
            dot.edge(arc.arc_tail.name, arc.arc_head.name, str(arc.arc_weight))
            dot_log.edge(arc.arc_tail.name, arc.arc_head.name, str(arc.log_arc_weight))

        dot.render(filename, view=True)
        dot_log.render((filename + "_logarithmic"), view=True)


class Graph():

    def __init__(self):
        self.set_of_nodes = []
        self.set_of_arcs = []
        self.arc_matrix = []
        self.number_of_nodes = 0
        self.table_header = []

    def detect_negative_cycles(self):
        distance_matrix = [[float('inf') for i in self.set_of_nodes] for j in self.set_of_nodes]
        predecessor_matrix = [[None for i in self.set_of_nodes] for j in self.set_of_nodes]

        # Initialization
        for i, node_i in enumerate(self.set_of_nodes):
            distance_matrix[i][i] = 0
            predecessor_matrix[i][i] = node_i

            for j, node_j in enumerate(self.set_of_nodes):
                if i == j:
                    pass
                else:
                    if self.arc_matrix[i][j]:
                        distance_matrix[i][j] = self.arc_matrix[i][j].log_arc_weight
                        predecessor_matrix[i][j] = node_i
                    else:
                        distance_matrix[i][j] = float('inf')
                        predecessor_matrix[i][j] = None

        # Iterations
        for v, node_v in enumerate(self.set_of_nodes):
            for i, node_i in enumerate(self.set_of_nodes):
                for j, node_j in enumerate(self.set_of_nodes):
                    if distance_matrix[i][j] > distance_matrix[i][v] + distance_matrix[v][j]:
                        distance_matrix[i][j] = distance_matrix[i][v] + distance_matrix[v][j]
                        predecessor_matrix[i][j] = predecessor_matrix[v][j]

        # Negative cycle?
        final_node = None
        predecessor_node = None

        for i, node_i in enumerate(self.set_of_nodes):
            if distance_matrix[i][i] < 0:
                final_node = node_i
                predecessor_node = predecessor_matrix[i][i]
                start_node_index = i
                break

        beginning_reached = False

        cyclic_path = []
        cyclic_path.append(final_node)
        cyclic_path.append(predecessor_node)

        while not beginning_reached:
            current_node_index = self.set_of_nodes.index(cyclic_path[-1])
            predecessor_node = predecessor_matrix[start_node_index][current_node_index]
            cyclic_path.append(predecessor_node)

            if cyclic_path[0] == cyclic_path[-1]:
                beginning_reached = True

        cyclic_path.reverse()

        print("")
        print("-- Distances Matrix --")
        print([str(_) for _ in self.table_header])
        print("-----------------------")
        for elem in distance_matrix:
            print(elem)

        print("")
        print("-- Predecessor Matrix --")
        print([str(_) for _ in self.table_header])
        print("-----------------------")
        for elem in predecessor_matrix:
            dummy = []
            for _ in elem:
                try:
                    dummy.append(_.name)
                except:
                    dummy.append("-")
            print([_ for _ in dummy])

        print("")
        print("-- Path for negative cycle --")
        print([_.name for _ in cyclic_path])


class Node():
    def __init__(self, name):
        self.name = str(name)


class Arc():
    def __init__(self, from_node, to_node, arc_weight):
        self.arc_weight = arc_weight
        self.arc_head = to_node
        self.arc_tail = from_node
        self.log_arc_weight = -math.log(float(self.arc_weight))


if __name__ == "__main__":
    My_Arbitrage_Project = ArbitrageProject()
    My_Arbitrage_Project.read_excel_file("arbitrage.xlsx")
    My_Arbitrage_Project.print_graph("arbitrage")
    My_Arbitrage_Project.graph.detect_negative_cycles()
